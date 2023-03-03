# 导入openpyxl模块
import openpyxl

# 创建工作簿和工作表
wb = openpyxl.Workbook()
ws = wb.active

# 创建字体对象
font = openpyxl.styles.Font(name='Arial', size=16, bold=True, color='FF0000')

# 创建对齐对象
alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

# 创建边框对象
border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin', color='000000'),
                                right=openpyxl.styles.Side(style='thin', color='000000'),
                                top=openpyxl.styles.Side(style='thin', color='000000'),
                                bottom=openpyxl.styles.Side(style='thin', color='000000'))

# 创建填充对象
fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='00FF00')

# 设置A1单元格的值和样式
ws['A1'] = 'Hello'
ws['A1'].font = font
ws['A1'].alignment = alignment
ws['A1'].border = border
ws['A1'].fill = fill

# 设置行高和列宽
ws.row_dimensions[1].height = 40 # 设置第一行的高度为40像素
ws.column_dimensions['A'].width = 20 # 设置第一列的宽度为20字符

# 合并B2到D4单元格，并设置值和样式
ws.merge_cells('B2:D4')
ws['B2'] = 'World'
ws['B2'].font = font
ws['B2'].alignment = alignment

# 保存工作簿为example.xlsx文件
wb.save('example.xlsx')

#范例
def save_xls(result,outname):
    info,OD = result
    for line in info:
        line[7] = '\n'.join(line[7])
        line[8] = '\n'.join(line[8])
    #info = list(map(list,zip(*info)))
    df = pd.DataFrame(info, columns=['班次','站点','到站时间','实际上车人数','预测上车人数','实际下车人数','预测下车人数','上车视频','下车视频'])
    writer = pd.ExcelWriter(outname+'.xlsx', engine='openpyxl')
    df.to_excel(writer, sheet_name='Sheet1', index=False)

    df2 = pd.DataFrame([OD], columns=["上车客流绝对精度","下车客流绝对精度","总客流绝对精度","上车客流精度","下车客流精度","总客流精度"])
    df2.to_excel(writer,sheet_name="Sheet1",startrow = df.shape[0]+4,index=False)

    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    worksheet.column_dimensions['H'].width = 30
    worksheet.column_dimensions['I'].width = 30
    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 15
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = alignment
    writer.save()